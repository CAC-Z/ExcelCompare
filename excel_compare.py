# 延迟导入重库（pandas），提升单文件启动速度
import importlib
pd = None  # runtime lazy import
from PyQt5 import QtWidgets, QtGui, QtCore
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QFormLayout, QScrollArea, QWidget, QGroupBox, QVBoxLayout, QHBoxLayout, QListWidget, QListWidgetItem, QPushButton, QLabel, QLineEdit, QGridLayout, QComboBox
from PyQt5.QtGui import QFont
from PyQt5.QtWidgets import QGraphicsDropShadowEffect
import sys
import os
from datetime import datetime


class FlowLayout(QtWidgets.QLayout):
    """简单的流式布局，使标签自动换行。"""
    def __init__(self, parent=None, margin=0, hspacing=8, vspacing=8):
        super().__init__(parent)
        self._hspacing = hspacing
        self._vspacing = vspacing
        self.itemList = []
        self.setContentsMargins(margin, margin, margin, margin)

    def addItem(self, item):
        self.itemList.append(item)

    def count(self):
        return len(self.itemList)

    def itemAt(self, index):
        if 0 <= index < len(self.itemList):
            return self.itemList[index]
        return None

    def takeAt(self, index):
        if 0 <= index < len(self.itemList):
            return self.itemList.pop(index)
        return None

    def expandingDirections(self):
        return QtCore.Qt.Orientations(QtCore.Qt.Orientation(0))

    def hasHeightForWidth(self):
        return True

    def heightForWidth(self, width):
        return self.doLayout(QtCore.QRect(0, 0, width, 0), True)

    def setGeometry(self, rect):
        super().setGeometry(rect)
        self.doLayout(rect, False)

    def sizeHint(self):
        return self.minimumSize()

    def minimumSize(self):
        size = QtCore.QSize()
        for item in self.itemList:
            size = size.expandedTo(item.minimumSize())
        mleft, mtop, mright, mbottom = self.getContentsMargins()
        size += QtCore.QSize(mleft + mright, mtop + mbottom)
        return size

    def doLayout(self, rect, testOnly):
        x = rect.x()
        y = rect.y()
        lineHeight = 0
        mleft, mtop, mright, mbottom = self.getContentsMargins()
        effectiveRect = rect.adjusted(mleft, mtop, -mright, -mbottom)
        x = effectiveRect.x()
        y = effectiveRect.y()
        maxWidth = effectiveRect.right()
        for item in self.itemList:
            wid = item.widget()
            spaceX = self._hspacing
            spaceY = self._vspacing
            nextX = x + item.sizeHint().width() + spaceX
            if nextX - spaceX > maxWidth and lineHeight > 0:
                x = effectiveRect.x()
                y = y + lineHeight + spaceY
                nextX = x + item.sizeHint().width() + spaceX
                lineHeight = 0
            if not testOnly:
                item.setGeometry(QtCore.QRect(QtCore.QPoint(x, y), item.sizeHint()))
            x = nextX
            lineHeight = max(lineHeight, item.sizeHint().height())
        return y + lineHeight - rect.y()

class CompareToolApp(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Excel 数据对比工具")
        # 获取程序运行的路径
        if getattr(sys, 'frozen', False):
            bundle_dir = sys._MEIPASS
        else:
            bundle_dir = os.path.dirname(os.path.abspath(__file__))

        icon_path = os.path.join(bundle_dir, 'icons', 'icon.png')
        self.setWindowIcon(QtGui.QIcon(icon_path))
        # 全局现代样式：卡片、窄滚动条
        self.setStyleSheet("""
            QWidget { font-size: 16px; }
            QGroupBox { 
                font-weight: bold; 
                margin-top: 12px; 
                border: 1px solid #e6e6e6; 
                border-radius: 10px; 
                background: #ffffff; 
                padding: 10px 12px 12px 12px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top left;
                left: 12px;
                padding: 0 6px;
                background: transparent;
            }
            QLabel[role="file"] { background: #f8f9fb; border: 1px solid #e6e8eb; border-radius: 6px; padding: 6px 8px; }
            QListWidget { border: 1px solid #e6e6e6; border-radius: 8px; }
            /* 统一输入框样式 */
            QLineEdit { border: 1px solid #e0e3e7; border-radius: 6px; padding: 6px 10px; background: #f8f9fb; }
            QLineEdit:hover { border-color: #c9ced6; background: #f7f9fc; }
            QLineEdit:focus { border-color: #7aa7ff; background: #ffffff; }
            /* 筛选输入框：图标由 addAction 提供，这里不再额外绘制，避免位置异常 */
            QLineEdit[role="filter"] { padding-left: 0px; }
            /* 下拉菜单现代样式 */
            QComboBox { border: 1px solid #e0e3e7; border-radius: 6px; padding: 6px 8px; background: #fff; }
            QComboBox:hover { border-color: #c9ced6; }
            /* 可编辑只读行编辑器外观，避免双重边框并让文本区域更自然 */
            QComboBox QLineEdit { border: none; padding: 0 8px; background: transparent; }
            QComboBox::drop-down { 
                subcontrol-origin: padding; 
                subcontrol-position: top right; 
                width: 28px; 
                border-left: 1px solid #e0e3e7; 
                background: transparent; 
            }
            QComboBox::down-arrow { image: url('icons/chevron-down.svg'); width: 12px; height: 12px; }
            QComboBox:hover::down-arrow { image: url('icons/chevron-down-dark.svg'); }
            /* 强调可点击的下拉：轻微悬停底色，仅对设置了 fullClick 的生效 */
            QComboBox[fullClick="true"]:hover { background: #f7faff; }
            QComboBox QAbstractItemView { border: 1px solid #e0e3e7; outline: none; background: #ffffff; }
            QComboBox QAbstractItemView::item { padding: 6px 8px; color: #1f2328; }
            QComboBox QAbstractItemView::item:hover { background: #f4f8ff; color: #1f2328; }
            QComboBox QAbstractItemView::item:selected { background: #e9f2ff; color: #1f2328; }
            QScrollBar:vertical {
                background: transparent;
                width: 8px;
                margin: 2px 2px 2px 2px;
            }
            QScrollBar::handle:vertical {
                background: #c2c7d0; border-radius: 4px; min-height: 20px;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                height: 0px; background: none; border: none;
            }
            QScrollBar:horizontal { height: 8px; }
            QScrollBar::handle:horizontal { background: #c2c7d0; border-radius: 4px; min-width: 20px;}
            QPushButton[cssClass="primary"] { background-color: #007BFF; color: white; padding: 10px 25px; border-radius: 8px; }
            QPushButton[cssClass="primary"]:hover { background-color: #0069d9; }
            QPushButton[cssClass="success"] { background-color: #28A745; color: white; padding: 12px 40px; border-radius: 8px; }
            QPushButton[cssClass="success"]:hover { background-color: #218838; }
            QPushButton[cssClass="ghost"] { background-color: #f5f6f7; color: #333; padding: 8px 14px; border-radius: 6px; border: 1px solid #e0e0e0; }
            QPushButton[cssClass="ghost"]:hover { background-color: #eceff3; }
            QLabel.badge { background: #f0f2f5; border: 1px solid #e0e0e0; border-radius: 14px; padding: 4px 10px; }
            /* QMessageBox 更紧凑的字体与按钮样式 */
            QMessageBox { font-size: 14px; }
            QMessageBox QLabel { font-size: 14px; color: #1f2328; }
            QMessageBox QPushButton { 
                font-size: 14px; 
                padding: 6px 14px; 
                border-radius: 6px; 
                background: #f5f6f7; 
                color: #333; 
                border: 1px solid #e0e0e0; 
                min-width: 72px;
            }
            QMessageBox QPushButton:hover { background: #eceff3; }
        """)

        # File paths and display names
        self.file1_path = None
        self.file2_path = None
        # UI显示名（截断后用于控件），以及完整名（用于导出/逻辑）
        self.file1_display_name_str_full = "文件1"
        self.file2_display_name_str_full = "文件2"
        self.file1_display_name_str = "文件1"
        self.file2_display_name_str = "文件2"


        # Main layout
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(15)

        # 顶部标题（按需显示，需求为移除，故省略）

        # 文件选择卡片：按钮 + 文件名行
        self.files_card = QGroupBox("文件选择")
        files_grid = QGridLayout()
        files_grid.setColumnStretch(0, 0)
        files_grid.setColumnStretch(1, 1)
        files_grid.setHorizontalSpacing(10)
        self.files_card.setLayout(files_grid)

        self.file1_button = QtWidgets.QPushButton("选择文件1")
        self.file1_button.setProperty('cssClass', 'primary')
        self.file1_button.clicked.connect(self.load_file1)
        self.file1_button.setFocusPolicy(QtCore.Qt.NoFocus)
        self.file1_button.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.file1_label = QtWidgets.QLabel("未选择文件")
        self.file1_label.setAlignment(QtCore.Qt.AlignLeft)
        self.file1_label.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Fixed)
        self.file1_label.setProperty('role', 'file')

        self.file2_button = QtWidgets.QPushButton("选择文件2")
        self.file2_button.setProperty('cssClass', 'primary')
        self.file2_button.clicked.connect(self.load_file2)
        self.file2_button.setFocusPolicy(QtCore.Qt.NoFocus)
        self.file2_button.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.file2_label = QtWidgets.QLabel("未选择文件")
        self.file2_label.setAlignment(QtCore.Qt.AlignLeft)
        self.file2_label.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Fixed)
        self.file2_label.setProperty('role', 'file')

        files_grid.addWidget(self.file1_button, 0, 0)
        files_grid.addWidget(self.file1_label, 0, 1)
        files_grid.addWidget(self.file2_button, 1, 0)
        files_grid.addWidget(self.file2_label, 1, 1)
        main_layout.addWidget(self.files_card)

        # 索引设置卡片（下拉选择）
        self.index_card = QGroupBox("索引设置")
        index_section_layout = QFormLayout()
        self.index_card.setLayout(index_section_layout)
        self.index1_label = QtWidgets.QLabel(f"{self.file1_display_name_str} 索引列：")
        self.index1_combo = QComboBox(); self.index1_combo.setEditable(True)
        # 使用可编辑 + 只读的行编辑器来显示占位文本，避免中间箭头问题
        self.index1_combo.lineEdit().setReadOnly(True)
        self.index1_combo.lineEdit().setPlaceholderText("请选择索引列")
        self.index1_combo.setProperty('fullClick', True)
        self.index1_combo.setProperty('popupOpen', False)
        # 手型指针（本体、编辑器、弹出列表）
        self.index1_combo.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.index1_combo.lineEdit().setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.index1_combo.setSizeAdjustPolicy(QComboBox.AdjustToContents)
        # 事件过滤：任意区域点击展开/收起
        self.index1_combo.installEventFilter(self)
        self.index1_combo.lineEdit().installEventFilter(self)
        # 下拉列表项也显示手型，并在选择后同步关闭状态
        try:
            self.index1_combo.view().setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
            self.index1_combo.view().viewport().setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        except Exception:
            pass
        self.index1_combo.activated.connect(lambda _=None, c=self.index1_combo: self._on_combo_close(c))
        self.index1_combo.currentIndexChanged.connect(lambda _=None, c=self.index1_combo: self._on_combo_close(c))
        index_section_layout.addRow(self.index1_label, self.index1_combo)
        self.index2_label = QtWidgets.QLabel(f"{self.file2_display_name_str} 索引列：")
        self.index2_combo = QComboBox(); self.index2_combo.setEditable(True)
        self.index2_combo.lineEdit().setReadOnly(True)
        self.index2_combo.lineEdit().setPlaceholderText("请选择索引列")
        self.index2_combo.setProperty('fullClick', True)
        self.index2_combo.setProperty('popupOpen', False)
        self.index2_combo.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.index2_combo.lineEdit().setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.index2_combo.setSizeAdjustPolicy(QComboBox.AdjustToContents)
        self.index2_combo.installEventFilter(self)
        self.index2_combo.lineEdit().installEventFilter(self)
        try:
            self.index2_combo.view().setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
            self.index2_combo.view().viewport().setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        except Exception:
            pass
        self.index2_combo.activated.connect(lambda _=None, c=self.index2_combo: self._on_combo_close(c))
        self.index2_combo.currentIndexChanged.connect(lambda _=None, c=self.index2_combo: self._on_combo_close(c))
        index_section_layout.addRow(self.index2_label, self.index2_combo)
        main_layout.addWidget(self.index_card)

        # 智能对比项设置卡片
        self.mapping_card = QGroupBox("对比项配置")
        mapping_card_layout = QVBoxLayout()
        self.mapping_card.setLayout(mapping_card_layout)

        # 顶部：左右列清单 + 过滤
        lists_row = QHBoxLayout()
        left_col = QVBoxLayout(); right_col = QVBoxLayout()
        self.left_filter = QLineEdit(); self.left_filter.setPlaceholderText("筛选文件1列名...")
        self.left_filter.setProperty('role', 'filter')
        self.left_filter.setClearButtonEnabled(True)
        self.right_filter = QLineEdit(); self.right_filter.setPlaceholderText("筛选文件2列名...")
        self.right_filter.setProperty('role', 'filter')
        self.right_filter.setClearButtonEnabled(True)
        # 占位符颜色更柔和
        pal1 = self.left_filter.palette(); pal1.setColor(QtGui.QPalette.PlaceholderText, QtGui.QColor('#98a2ad')); self.left_filter.setPalette(pal1)
        pal2 = self.right_filter.palette(); pal2.setColor(QtGui.QPalette.PlaceholderText, QtGui.QColor('#98a2ad')); self.right_filter.setPalette(pal2)
        # 使用 addAction 添加左侧放大镜，位置稳定，不会跑到文本右上角
        try:
            search_icon_path = os.path.join(bundle_dir, 'icons', 'search.svg')
            self.left_filter.addAction(QtGui.QIcon(search_icon_path), QLineEdit.LeadingPosition)
            self.right_filter.addAction(QtGui.QIcon(search_icon_path), QLineEdit.LeadingPosition)
        except Exception:
            pass
        self.left_list = QListWidget(); self.right_list = QListWidget()
        # 列表可点击：手型光标
        self.left_list.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.left_list.viewport().setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.right_list.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.right_list.viewport().setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.left_list.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
        self.right_list.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
        left_col.addWidget(self.left_filter); left_col.addWidget(self.left_list)
        right_col.addWidget(self.right_filter); right_col.addWidget(self.right_list)
        # 默认显示约8行（随字体自适应）
        fm = self.left_list.fontMetrics()
        row_h = fm.height() + 6
        target_h = row_h * 8 + 8
        self.left_list.setFixedHeight(target_h)
        self.right_list.setFixedHeight(target_h)
        lists_row.addLayout(left_col); lists_row.addLayout(right_col)

        # 中部：动作按钮
        actions_row = QHBoxLayout()
        self.btn_add_pair = QPushButton("添加对比项 ▶")
        self.btn_add_pair.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.btn_add_pair.setFocusPolicy(QtCore.Qt.NoFocus)
        self.btn_add_pair.setProperty('cssClass', 'ghost')
        self.btn_auto_pair = QPushButton("自动匹配同名列")
        self.btn_auto_pair.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.btn_auto_pair.setFocusPolicy(QtCore.Qt.NoFocus)
        self.btn_auto_pair.setProperty('cssClass', 'ghost')
        actions_row.addStretch(1)
        actions_row.addWidget(self.btn_auto_pair)
        actions_row.addWidget(self.btn_add_pair)

        # 底部：已选对比项标签区（放入可滚动容器以限制高度，避免挤压整体布局）
        self.tags_container = QWidget(); self.tags_layout = FlowLayout(hspacing=8, vspacing=6)
        self.tags_layout.setContentsMargins(8, 8, 8, 8)
        self.tags_container.setLayout(self.tags_layout)
        self.tags_scroll = QScrollArea(); self.tags_scroll.setWidgetResizable(True)
        self.tags_scroll.setWidget(self.tags_container)
        self.tags_scroll.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.tags_scroll.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
        self.tags_scroll.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
        # 提升可视空间，默认可展示更多行标签
        self.tags_scroll.setMaximumHeight(200)

        mapping_card_layout.addLayout(lists_row)
        mapping_card_layout.addLayout(actions_row)
        mapping_card_layout.addWidget(self.tags_scroll)
        main_layout.addWidget(self.mapping_card)

        # 保存对比映射的数据结构
        self.mappings = []  # list of dicts {col1, col2}

        # 废弃旧的手工输入区域（保留结构以防后续扩展），现由标签配置替代

        # Compare button
        self.compare_button = QtWidgets.QPushButton("开始对比")
        self.compare_button.setProperty('cssClass', 'success')
        self.compare_button.clicked.connect(self.compare_files)
        # 移除按钮虚线焦点框
        self.compare_button.setFocusPolicy(QtCore.Qt.NoFocus)
        self.compare_button.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        main_layout.addWidget(self.compare_button, alignment=QtCore.Qt.AlignCenter)

        # Status label
        self.status_label = QtWidgets.QLabel("准备就绪")
        self.status_label.setAlignment(QtCore.Qt.AlignCenter)
        self.status_label.setStyleSheet("font-weight: bold;")
        main_layout.addWidget(self.status_label)

        self.setLayout(main_layout)
        self.setMinimumSize(640, 760)
        self.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        self.resize(640, 800)
        # 应用卡片阴影
        self.apply_card_shadow(self.files_card)
        self.apply_card_shadow(self.index_card)
        self.apply_card_shadow(self.mapping_card)

    def load_file1(self):
        options = QFileDialog.Options()
        file1, _ = QFileDialog.getOpenFileName(self, "选择文件1", "", "Excel files (*.xlsx);;All Files (*)", options=options)
        if file1:
            self.file1_path = file1
            base_name = os.path.basename(self.file1_path)
            # 完整名与截断名（用于UI显示）
            self.file1_display_name_str_full = os.path.splitext(base_name)[0]
            self.file1_display_name_str = self._truncate_ui_name(self.file1_display_name_str_full)
            # 文件标签显示带扩展名的截断文本，并提供完整toolTip
            self.file1_label.setText(self._truncate_ui_name(base_name, 32))
            self.file1_label.setToolTip(base_name)
            self.update_all_labels()
            self.refresh_column_lists()

    def load_file2(self):
        options = QFileDialog.Options()
        file2, _ = QFileDialog.getOpenFileName(self, "选择文件2", "", "Excel files (*.xlsx);;All Files (*)", options=options)
        if file2:
            self.file2_path = file2
            base_name = os.path.basename(self.file2_path)
            # 完整名与截断名（用于UI显示）
            self.file2_display_name_str_full = os.path.splitext(base_name)[0]
            self.file2_display_name_str = self._truncate_ui_name(self.file2_display_name_str_full)
            # 文件标签显示带扩展名的截断文本，并提供完整toolTip
            self.file2_label.setText(self._truncate_ui_name(base_name, 32))
            self.file2_label.setToolTip(base_name)
            self.update_all_labels()
            self.refresh_column_lists()
            
    def _truncate_ui_name(self, text: str, max_len: int = 20) -> str:
        """将文本按字符长度中间截断，避免前端控件被拉伸。"""
        if text is None:
            return ""
        if len(text) <= max_len:
            return text
        keep = max_len - 3
        left = keep // 2
        right = keep - left
        return f"{text[:left]}...{text[-right:]}"

    def update_all_labels(self):
        """统一更新所有与文件名相关的标签"""
        # 使用截断名用于界面，保持紧凑
        self.index1_label.setText(f"{self.file1_display_name_str} 索引列：")
        self.index2_label.setText(f"{self.file2_display_name_str} 索引列：")
        # 标签式配置无需手填列数

    def apply_card_shadow(self, gb: QWidget):
        effect = QGraphicsDropShadowEffect(self)
        effect.setBlurRadius(16)
        effect.setOffset(0, 2)
        effect.setColor(QtGui.QColor(0, 0, 0, 40))
        gb.setGraphicsEffect(effect)

    def _read_excel_header_fast(self, file_path: str):
        """仅读取 Excel 第一行作为列名，避免导入 pandas，提升启动与选择速度。"""
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            wb.close()
            cols = []
            for c in row:
                if c is None:
                    continue
                text = str(c).strip()
                if text:
                    cols.append(text)
            return cols
        except Exception:
            return []

    # 让 QComboBox 在任意区域点击都可展开/再次点击收起
    def eventFilter(self, obj, event):
        try:
            if event.type() == QtCore.QEvent.MouseButtonPress:
                # 点击 lineEdit：控制父级 QComboBox 的弹出/收起
                if isinstance(obj, QtWidgets.QLineEdit) and isinstance(obj.parent(), QtWidgets.QComboBox):
                    combo = obj.parent()
                    if combo.property('fullClick'):
                        if combo.property('popupOpen'):
                            combo.hidePopup()
                            combo.setProperty('popupOpen', False)
                        else:
                            combo.showPopup()
                            combo.setProperty('popupOpen', True)
                        return True
                # 点击 QComboBox 本体
                if isinstance(obj, QtWidgets.QComboBox) and obj.property('fullClick'):
                    if obj.property('popupOpen'):
                        obj.hidePopup()
                        obj.setProperty('popupOpen', False)
                    else:
                        obj.showPopup()
                        obj.setProperty('popupOpen', True)
                    return True
        except Exception:
            pass
        return super().eventFilter(obj, event)

    def _on_combo_close(self, combo: QtWidgets.QComboBox):
        combo.setProperty('popupOpen', False)


    # 新：读取两边列名并填充到列表，便于点选/筛选
    def refresh_column_lists(self):
        self.left_list.clear(); self.right_list.clear()

        # 读取两边列名，仅表头
        self.cols1, self.cols2 = [], []
        try:
            if self.file1_path:
                self.cols1 = self._read_excel_header_fast(self.file1_path)
            if self.file2_path:
                self.cols2 = self._read_excel_header_fast(self.file2_path)
        except Exception:
            pass

        # 更新索引下拉
        prev1 = self.index1_combo.currentText() if hasattr(self, 'index1_combo') else ''
        prev2 = self.index2_combo.currentText() if hasattr(self, 'index2_combo') else ''
        if hasattr(self, 'index1_combo'):
            try:
                self.index1_combo.currentTextChanged.disconnect()
            except Exception:
                pass
            self.index1_combo.blockSignals(True)
            self.index1_combo.clear()
            self.index1_combo.addItems(self.cols1)
            if prev1 in self.cols1:
                self.index1_combo.setCurrentText(prev1)
            else:
                # 未选中时显示占位文本，避免内容区出现额外箭头
                self.index1_combo.setCurrentIndex(-1)
            self.index1_combo.blockSignals(False)
            self.index1_combo.currentTextChanged.connect(self.on_index_changed)
        if hasattr(self, 'index2_combo'):
            try:
                self.index2_combo.currentTextChanged.disconnect()
            except Exception:
                pass
            self.index2_combo.blockSignals(True)
            self.index2_combo.clear()
            self.index2_combo.addItems(self.cols2)
            if prev2 in self.cols2:
                self.index2_combo.setCurrentText(prev2)
            else:
                self.index2_combo.setCurrentIndex(-1)
            self.index2_combo.blockSignals(False)
            self.index2_combo.currentTextChanged.connect(self.on_index_changed)

        # 选中的索引列
        idx1 = self.index1_combo.currentText() if self.index1_combo.count() else None
        idx2 = self.index2_combo.currentText() if self.index2_combo.count() else None

        # 根据索引排除对比候选列
        left_cols = [c for c in self.cols1 if c and c != idx1]
        right_cols = [c for c in self.cols2 if c and c != idx2]

        # 填充候选列表
        for c in left_cols:
            self.left_list.addItem(QListWidgetItem(c))
        for c in right_cols:
            self.right_list.addItem(QListWidgetItem(c))

        # 移除包含索引列的已选映射
        if getattr(self, 'mappings', None) is not None:
            self.mappings = [m for m in self.mappings if m.get('col1') not in (idx1,) and m.get('col2') not in (idx2,)]

        # 绑定过滤（重连）
        try:
            self.left_filter.textChanged.disconnect()
        except Exception:
            pass
        try:
            self.right_filter.textChanged.disconnect()
        except Exception:
            pass
        def apply_filter(list_widget: QListWidget, text: str, side: str):
            t = (text or '').strip().lower()
            used_left = {m['col1'] for m in getattr(self, 'mappings', [])}
            used_right = {m['col2'] for m in getattr(self, 'mappings', [])}
            for i in range(list_widget.count()):
                it = list_widget.item(i)
                hide_by_text = t not in it.text().lower()
                hide_by_used = it.text() in (used_left if side == 'left' else used_right)
                it.setHidden(hide_by_text or hide_by_used)
        self.left_filter.textChanged.connect(lambda v: apply_filter(self.left_list, v, 'left'))
        self.right_filter.textChanged.connect(lambda v: apply_filter(self.right_list, v, 'right'))

        # 绑定动作按钮
        try:
            self.btn_add_pair.clicked.disconnect()
            self.btn_auto_pair.clicked.disconnect()
        except Exception:
            pass
        self.btn_add_pair.clicked.connect(self.add_pair_from_selection)
        self.btn_auto_pair.clicked.connect(self.auto_pair_by_same_name)

        # 重绘标签 + 初次过滤（应用搜索词与已选映射隐藏）
        self.render_tags()
        apply_filter(self.left_list, self.left_filter.text(), 'left')
        apply_filter(self.right_list, self.right_filter.text(), 'right')

    def on_index_changed(self, _):
        # 索引变更后刷新候选列
        self.refresh_column_lists()

    def add_pair_from_selection(self):
        left = self.left_list.currentItem().text() if self.left_list.currentItem() else None
        right = self.right_list.currentItem().text() if self.right_list.currentItem() else None
        if not left or not right:
            QMessageBox.information(self, "提示", "请在左右列清单中各选择一项后再添加。")
            return
        if any(m['col1'] == left and m['col2'] == right for m in self.mappings):
            return
        self.mappings.append({'col1': left, 'col2': right})
        self.render_tags()
        # 隐藏已加入的候选项，并清除选择
        self.apply_candidate_filters()
        self.left_list.clearSelection(); self.right_list.clearSelection()

    def auto_pair_by_same_name(self):
        left_names = {self.left_list.item(i).text() for i in range(self.left_list.count())}
        right_names = {self.right_list.item(i).text() for i in range(self.right_list.count())}
        commons = sorted(left_names.intersection(right_names))
        added = 0
        for name in commons:
            if not any(m['col1'] == name and m['col2'] == name for m in self.mappings):
                self.mappings.append({'col1': name, 'col2': name})
                added += 1
        if added == 0:
            QMessageBox.information(self, "提示", "未发现可自动匹配的同名列。")
        self.render_tags()
        self.apply_candidate_filters()

    def remove_tag(self, idx: int):
        if 0 <= idx < len(self.mappings):
            del self.mappings[idx]
            self.render_tags()
            self.apply_candidate_filters()

    def render_tags(self):
        # 清空标签区域
        while self.tags_layout.count():
            it = self.tags_layout.takeAt(0)
            if it:
                w = it.widget()
                if w:
                    w.deleteLater()

        # 渲染每个映射为一个“小芯片”，简化边框层级
        for i, m in enumerate(self.mappings):
            chip = QWidget()
            h = QHBoxLayout(); h.setContentsMargins(12, 6, 8, 6); h.setSpacing(8)
            chip.setLayout(h)
            # 单层外框 + 轻微阴影，无内部小框
            chip.setStyleSheet("QWidget { border: 1px solid #e3e7ec; border-radius: 14px; background: #f6f8fb; }")
            label = QLabel(f"{m['col1']}  ⇄  {m['col2']}")
            # 关闭删除按钮的边框与底色，仅保留字符
            btn = QPushButton("✕")
            btn.setFixedSize(20, 20)
            btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
            btn.setFocusPolicy(QtCore.Qt.NoFocus)
            btn.setStyleSheet("QPushButton { border: none; background: transparent; color: #9aa3ad; font-weight: bold; } QPushButton:hover { color: #4a5560; }")
            btn.clicked.connect(lambda _, idx=i: self.remove_tag(idx))
            h.addWidget(label)
            h.addWidget(btn)
            self.tags_layout.addWidget(chip)
        # FlowLayout 无需添加 stretch

    def apply_candidate_filters(self):
        """根据搜索文本与已选映射，隐藏候选列表中已加入映射的列。"""
        try:
            t1 = self.left_filter.text() if hasattr(self, 'left_filter') else ''
            t2 = self.right_filter.text() if hasattr(self, 'right_filter') else ''
            used_left = {m['col1'] for m in getattr(self, 'mappings', [])}
            used_right = {m['col2'] for m in getattr(self, 'mappings', [])}
            for i in range(self.left_list.count()):
                it = self.left_list.item(i)
                hide_by_text = (t1 or '').strip().lower() not in it.text().lower()
                hide_by_used = it.text() in used_left
                it.setHidden(hide_by_text or hide_by_used)
            for i in range(self.right_list.count()):
                it = self.right_list.item(i)
                hide_by_text = (t2 or '').strip().lower() not in it.text().lower()
                hide_by_used = it.text() in used_right
                it.setHidden(hide_by_text or hide_by_used)
        except Exception:
            pass


    def compare_files(self):
        if not self.file1_path or not self.file2_path:
            QMessageBox.critical(self, "错误", "请先选择两个文件")
            return

        index1 = self.index1_combo.currentText().strip() if self.index1_combo.count() else ''
        index2 = self.index2_combo.currentText().strip() if self.index2_combo.count() else ''
        if not index1 or not index2:
            QMessageBox.critical(self, "错误", "请填写索引列")
            return

        try:
            global pd
            if pd is None:
                # 延迟导入 pandas，避免影响启动
                pd = importlib.import_module('pandas')
            self.status_label.setText("正在读取文件...")
            df1 = pd.read_excel(self.file1_path, header=0, engine='openpyxl')
            df2 = pd.read_excel(self.file2_path, header=0, engine='openpyxl')

            if index1 not in df1.columns or index2 not in df2.columns:
                QMessageBox.critical(self, "错误", "找不到指定的索引列，请检查列名是否正确")
                return
            
            # 导出与结果展示使用完整文件名，避免被截断
            file1_name = getattr(self, 'file1_display_name_str_full', self.file1_display_name_str)
            file2_name = getattr(self, 'file2_display_name_str_full', self.file2_display_name_str)


            self.status_label.setText("正在处理重复值...")
            duplicates_df1 = df1[df1.duplicated(index1, keep=False)].copy()
            duplicates_df1.loc[:, '来源'] = file1_name
            duplicates_df2 = df2[df2.duplicated(index2, keep=False)].copy()
            duplicates_df2.loc[:, '来源'] = file2_name
            all_duplicates = pd.concat([duplicates_df1, duplicates_df2])
            if not all_duplicates.empty:
                dup_ts = datetime.now().strftime('%Y%m%d_%H%M%S')
                dup_base = f"两个表格中重复的名字_{dup_ts}"
                dup_filename = f"{dup_base}.xlsx"
                if os.path.exists(dup_filename):
                    n = 1
                    while os.path.exists(f"{dup_base}_{n}.xlsx"):
                        n += 1
                    dup_filename = f"{dup_base}_{n}.xlsx"
                all_duplicates.to_excel(dup_filename, index=False)
                self.status_label.setText(f"已导出重复索引：{dup_filename}")
            
            df1 = df1.drop_duplicates(subset=[index1], keep='first').set_index(index1)
            df2 = df2.drop_duplicates(subset=[index2], keep='first').set_index(index2)

            common_index = df1.index.intersection(df2.index)
            df1_common = df1.loc[common_index]
            df2_common = df2.loc[common_index]

            self.status_label.setText("正在对比数据...")
            
            # 从标签式配置读取映射，并过滤无效列
            column_mappings = []
            for m in getattr(self, 'mappings', []):
                col1 = str(m.get('col1', '')).strip()
                col2 = str(m.get('col2', '')).strip()
                if col1 and col2 and col1 in df1_common.columns and col2 in df2_common.columns:
                    column_mappings.append({'col1': col1, 'col2': col2})

            if not column_mappings:
                QMessageBox.warning(self, "注意", "没有有效的列进行对比。请检查您是否已填写对比列，以及列名是否正确。")
                self.status_label.setText("准备就绪")
                return

            # 统一归一化函数：去空白、去千分位、数字转为一致格式（避免 0 与 0.0 误判）
            def _normalize_series(s: 'pd.Series') -> 'pd.Series':
                # 缺失值占位，避免与空字符串混淆
                s = s.copy()
                s = s.where(~s.isna(), other="__MISSING__")
                # 转字符串并去除首尾空白
                s_str = s.astype(str).str.strip()
                # 去除可能的千分位逗号
                s_clean = s_str.str.replace(',', '', regex=False)
                # 能转数字的统一为数字格式，再转为字符串，去除多余的0和小数点
                s_num = pd.to_numeric(s_clean, errors='coerce')
                result = s_clean.copy()
                mask = s_num.notna()
                # 使用通用格式，最多15位有效数字，避免 1.0 与 1、以及 1.2300 与 1.23 的差异
                result.loc[mask] = s_num.loc[mask].map(lambda x: f"{x:.15g}")
                return result

            overall_mismatch_mask = pd.Series(False, index=df1_common.index)
            for mapping in column_mappings:
                s1 = _normalize_series(df1_common[mapping['col1']])
                s2 = _normalize_series(df2_common[mapping['col2']])
                overall_mismatch_mask |= (s1 != s2)
            
            mismatch_indices = df1_common.index[overall_mismatch_mask]

            if mismatch_indices.empty:
                QMessageBox.information(self, "完成", "所有对比列的数据完全一致！")
                self.status_label.setText("未发现不匹配项")
                return

            mismatched_df1 = df1_common.loc[mismatch_indices]
            mismatched_df2 = df2_common.loc[mismatch_indices]
            
            # 仅保留存在差异的列与行，避免把相同数据一并导出
            detailed_result_list = []
            for mapping in column_mappings:
                col1, col2 = mapping['col1'], mapping['col2']

                # 针对每一对列单独计算差异掩码（使用归一化后的值）
                pair_mask = (
                    _normalize_series(df1_common[col1])
                    != _normalize_series(df2_common[col2])
                )
                pair_indices = df1_common.index[pair_mask]
                if pair_indices.empty:
                    continue

                df1_subset = (
                    df1_common.loc[pair_indices, [col1]]
                    .rename(columns={col1: f"{file1_name}_{col1}"})
                )
                df2_subset = (
                    df2_common.loc[pair_indices, [col2]]
                    .rename(columns={col2: f"{file2_name}_{col2}"})
                )
                detailed_result_list.extend([df1_subset, df2_subset])

            # 若无任何差异对，构造一个空表；否则按索引对齐横向拼接
            if detailed_result_list:
                detailed_df = pd.concat(detailed_result_list, axis=1)
            else:
                detailed_df = pd.DataFrame(index=mismatch_indices)

            summary_records = []
            for index_val in mismatch_indices:
                for mapping in column_mappings:
                    col1, col2 = mapping['col1'], mapping['col2']
                    val1 = df1_common.loc[index_val, col1]
                    val2 = df2_common.loc[index_val, col2]

                    # 使用与总体一致的归一化比较，避免 0 与 0.0、空格等导致的误报
                    n1 = _normalize_series(pd.Series([val1])).iloc[0]
                    n2 = _normalize_series(pd.Series([val2])).iloc[0]
                    if n1 != n2:
                        summary_records.append({
                            index1: index_val,
                            '不一致的列': f"{col1} vs {col2}",
                            f'{file1_name}的值': val1,
                            f'{file2_name}的值': val2,
                        })
            summary_df = pd.DataFrame(summary_records)

            # 使用时间戳命名，避免覆盖；若重名则追加计数后缀
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            base = f"对比的结果_{ts}"
            output_filename = f"{base}.xlsx"
            if os.path.exists(output_filename):
                n = 1
                while os.path.exists(f"{base}_{n}.xlsx"):
                    n += 1
                output_filename = f"{base}_{n}.xlsx"
            with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
                summary_df.to_excel(writer, sheet_name='差异汇总', index=False)
                detailed_df.to_excel(writer, sheet_name='详细对比数据', index=True)
            
            QMessageBox.information(self, "完成", f"对比完成！结果已保存到 '{output_filename}'。\n\n"
                                               "文件中包含两个Sheet：\n"
                                               "1. 差异汇总：清晰列出每一项不同。\n"
                                               "2. 详细对比数据：并排展示所有差异行的数据。")
            self.status_label.setText("对比完成")

        except Exception as e:
            QMessageBox.critical(self, "错误", f"对比过程中发生错误: {e}")
            self.status_label.setText("对比过程中发生错误")

if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    window = CompareToolApp()
    window.show()
    sys.exit(app.exec_())
